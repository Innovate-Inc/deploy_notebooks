# %%

import jupytext
import re
import os
import sys
import time
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.cell import get_column_letter
from datetime import date, timedelta, datetime as dt
from timeit import default_timer as timer
from arcgis.features import FeatureLayer, Feature, FeatureSet
from arcgis.geometry import Geometry, Polygon, Point, union
from arcgis.geometry.filters import intersects
from arcgis.geometry.functions import buffer, project
from io import BytesIO
from uuid import uuid4
import base64
from arcgis.gis import GIS
import urllib
import requests
import json
from copy import deepcopy
from tenacity import retry, stop_after_attempt, after_log
import logging


# %%
# helper functions
########################################################################################################################
def get_logger():
    logging.basicConfig(stream=sys.stderr, level=logging.DEBUG)
    logger = logging.getLogger(__name__)
    return logger


@retry(stop=stop_after_attempt(3), after=after_log(get_logger(), logging.DEBUG))
def load_feature_set(url, where, orderby=None, recordcount=None, fields="*", returngeometry='true'):
    fl = FeatureLayer(url)
    try:
        results = fl.query(where=where, return_geometry=returngeometry, out_fields=fields, orderby=orderby,
                           return_all_records=True)
        for f in results.features:
            f.geometry['spatialReference'] = results.spatial_reference
            f.geometry = Geometry(f.geometry)

        #         my_map.draw(results.features)
        print('load_feature_set' + " " + where)
        # print(results)
        return results.features
    #         return feature_set
    except Exception as e:
        print(e)
        return None


@retry(stop=stop_after_attempt(3), after=after_log(get_logger(), logging.DEBUG))
def load_features_json(url, where=None, geometry=None, order_by=None, fields="*", attempts=2, auth_token=None,
                       out_wkid=4326):
    print(url)
    print(where)
    if geometry is not None:
        print('geometry')
    try:
        wkid = 4326
        # print('loading fs, attempt number 1')
        params = {'outFields': '*', 'f': 'json', 'inSR': wkid, 'ourSR': wkid,
                  'returnGeometry': True,
                  }
        if where:
            params["where"] = where
        else:
            params["where"] = "1=1"
        if geometry:
            params["geometry"] = geometry
            params['geometryType'] = 'esriGeometryPolygon'
            params['spatialRel'] = 'esriSpatialRelIntersects'
        if auth_token:
            params['token'] = auth_token

        data = urllib.parse.urlencode(params)
        if not url.endswith('/query'):
            url += '/query'
        request_response = requests.post(url, data=data,
                                         headers={'Content-Type': 'application/x-www-form-urlencoded'})
        if request_response.status_code != 200:
            raise Exception(f'request response status {request_response.status_code}')
        if request_response.json().get('error', None) is not None:
            raise Exception(request_response.json().get('error'))

        features = request_response.json().get('features', [])
        for f in features:
            if f.get('geometry', None) is None:
                print(f)
            f['geometry']['spatialReference'] = {'wkid': out_wkid}
        return features
    except Exception as e:
        print(e)
        print(url)
        attempts -= 1
        pass
        if attempts > 0:
            print(f'attempts left {attempts}')
            time.sleep(.1)
            load_features_json(url, where, geometry, order_by, fields, attempts)
        raise Exception(e)


@retry(stop=stop_after_attempt(3), after=after_log(get_logger(), logging.DEBUG))
def update_feature(input_feature, target_url, feature_id=None, id_field='GlobalID', attachment=None):
    target_fl = FeatureLayer(target_url)
    attributes = {}
    for field in target_fl.properties['fields']:
        attributes[field['name']] = input_feature.attributes.get(field['name'], None)
    new_feat = Feature(geometry=input_feature.geometry, attributes=attributes)
    if feature_id is not None:
        feature_id = format_global(feature_id)
        query = f"{id_field} = '{feature_id}'"
        existing_feat = target_fl.query(where=query, out_fields='*')
        if len(existing_feat.features) > 0:
            feature_attributes = existing_feat.features[0].attributes
            new_feat.attributes['OBJECTID'] = feature_attributes['OBJECTID']
            new_feat.attributes['GlobalID'] = feature_attributes['GlobalID']
            edit_features = {'updates': [new_feat]}
        else:
            print('Could not find matching feature to update')
            new_feat.attributes.pop('OBJECTID')
            new_feat.attributes.pop('GlobalID')
            new_feat.attributes['GlobalID'] = str(uuid4())
            new_feat.attributes[id_field] = feature_id
            edit_features = {'adds': [new_feat]}
            # update_feature(input_feature, target_url, feature_id=None, id_field=id_field, attachment=attachment)
    else:
        new_feat.attributes.pop('OBJECTID')
        new_feat.attributes.pop('GlobalID')
        new_feat.attributes['GlobalID'] = str(uuid4())

        edit_features = {'adds': [new_feat]}
    if attachment is not None:
        attachment['parentGlobalId'] = format_global(new_feat.attributes['GlobalID'])
        attachment['globalId'] = str(uuid4())

        r = target_fl.edit_features(**edit_features, use_global_ids=True, attachments={"adds": [attachment]})
    else:
        r = target_fl.edit_features(**edit_features, use_global_ids=True)
    # print(r)
    return r


@retry(stop=stop_after_attempt(3), after=after_log(get_logger(), logging.DEBUG))
def get_intersect(service_url, input_geom, token):
    try:
        if service_url[-6:] != '/query':
            service_url += '/query'
        print(service_url)
        geom_type = 'esriGeometryPoint' if input_geom.type == 'Point' else 'esriGeometryPolyline' if input_geom.type == 'Polyline' else 'esriGeometryPolygon'

        data = {'geometry': input_geom, 'f': 'json', 'outFields': '*', 'spatialRel': 'esriSpatialRelIntersects',
                'geometryType': geom_type, 'token': token}
        data = urllib.parse.urlencode(data)
        headers = {
            'content-type': 'application/x-www-form-urlencoded'
        }
        response = requests.post(service_url, data=data, headers=headers)
        results = response.json()
        if results.get('features') and len(results.get('features')) > 0:
            print(len(results.get('features')))
            return FeatureSet(features=results.get('features'))
        return FeatureSet(features=[])
    except Exception as e:
        print('get_intersect error: ')
        print(e)
        return None


@retry(stop=stop_after_attempt(3), after=after_log(get_logger(), logging.DEBUG))
def buffer_miles(geom, distance=10, unit=9030, out_wkid=3857, wkid_type='latestWkid'):
    if not isinstance(geom, Geometry):
        geom = Geometry(geom)
    in_wkid = geom.spatial_reference.get('wkid')
    buffer_geom = buffer([geom], in_sr=in_wkid, distances=distance,
                         unit=unit,
                         out_sr=out_wkid)[0]
    buffer_geom.spatialReference = {wkid_type: out_wkid}
    return buffer_geom


def format_global(input_id, braces=True):
    # id = id.lower()
    if braces:
        if '{' not in input_id:
            input_id = '{' + '{0}'.format(input_id) + '}'
        return input_id
    else:
        return input_id.replace('{', '').replace('}', '')


def populate_sheet(workbook, sheet, lyr_details: dict, fire_name, layer):
    header_fill = PatternFill(start_color='344C67', end_color='344C67', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    alternate_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
    ws = workbook.worksheets[sheet]
    tab_title = ws.cell(1, 1).value
    ws.cell(1, 1, tab_title.replace('[FireName]', fire_name))
    date_cell = ws.cell(3, 1)
    date_cell.value = f"Current as of: {lyr_details['update_date']}"
    source_cell = ws.cell(4, 1)
    source_cell.value = f"Data Source: {lyr_details['name']}"

    if layer.features:
        unwanted_fields = ['OBJECTID', 'SHAPE', 'AUTOID']
        fields = [f for f in layer.fields if not any(y in f['name'].upper() for y in unwanted_fields)]

        field_list = [f['name'] for f in fields]
        field_alias = [f['alias'] for f in fields]

        # reorder 'NAME' field to index 2
        if 'NAME' in field_alias:
            name_index = field_alias.index('NAME')
            ualias = field_alias[name_index]
            uname = field_list[name_index]
            field_list.pop(name_index)
            field_alias.pop(name_index)
            field_list.insert(2, uname)
            field_alias.insert(2, ualias)

        header_row = 5  # not zero based
        # add header - field aliases
        for header_cell, v in enumerate(field_alias, 1):
            _cell = ws.cell(header_row, header_cell, v)
            _cell.fill = header_fill
            _cell.font = header_font
            column_width = 4
            if len(v) > column_width:
                column_width = len(v)
                # print(str(len(v)) + '  cell:' +get_column_letter(header_cell))
            ws.column_dimensions[get_column_letter(header_cell)].width = column_width + 3
        # need to refactor column width to dictionary object. get max column width of data, maybe not to exceed 120
        # make sure the column is as least as wide as the column heading
        #
        for row_index, row in enumerate(layer.features, 6):  # start at row 6
            for cell, f in enumerate(field_list):
                _cell = ws.cell(row_index, cell + 1, row.attributes[f])
                if row_index % 2 != 0:
                    _cell.fill = alternate_fill


def removeRows(target_lyr, deletes, target_field):
    # target_lyr.features = [x for x in target_lyr.features if x.attributes[target_field] not in deletes]
    for f in target_lyr.features:
        if f.attributes[target_field] in deletes:
            print(f'removing feature: {target_field}={f.attributes[target_field]}')
            target_lyr.features.remove(f)
    for f in target_lyr.features:
        if f.attributes[target_field] in deletes:
            print(f'removing feature: {target_field}={f.attributes[target_field]}')
            target_lyr.features.remove(f)
    return target_lyr


def generate_fire_report(irwin_id, perimeter_id, token, config_settings: dict):
    if irwin_id:
        print("irwin_id: {}".format(irwin_id))
    elif perimeter_id:
        print("perimeter_id: {}".format(perimeter_id))
    startmain = timer()
    reportstatus = 'not generated'
    try:
        # path to report template
        gis = GIS('home')
        report_template_item = gis.content.get(config_settings['REPORT_TEMPLATE_ID'])
        report_template = report_template_item.download()
        # target_data = 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/arcgis/rest/services/R9_Fire_Perimeter_Buffers/FeatureServer/0'

        response = {'facilities': {'Total': 0}}
        if not irwin_id and not perimeter_id:
            return None
        if irwin_id:
            # query irwin layer
            fire_incident = get_irwin_info(config_settings['IRWIN_SERVICE'],
                                           where_statement=f"{config_settings['IRWIN_ID_FIELD']} = '{irwin_id}'")
            # fire_incident = FireIncident(irwinID=irwin_id)
            response['IRWINID'] = irwin_id
        if perimeter_id:
            fire_incident = get_perimeters(config_settings['PERIMETER_SERVICE'],
                                           where=f"{config_settings['PERIMETER_ID_FIELD']} = '{perimeter_id}'")
            # fire_incident = FireIncident(perimeterID=perimeter_id)
            response['perimeter_id'] = perimeter_id
            if fire_incident and fire_incident[0]['attributes'].get(config_settings['PERIMETER_IRWIN_FIELD']):
                response['IRWINID'] = fire_incident[0]['attributes'].get(config_settings['PERIMETER_IRWIN_FIELD'])
        # Error if fire not found
        if not fire_incident:
            raise Exception("fire not found")
        else:
            fire_incident = fire_incident[0]
            fire_incident['geometry'] = Geometry(fire_incident['geometry'])
            # fire_incident['geometry'] = Geometry(fire_incident['geometry'])
        # return buffer extent
        fire_incident['buffer'] = buffer_miles(fire_incident['geometry'], distance=2000)
        response['FireBufferExtent'] = get_extent(fire_incident['buffer'])
        # fire feature id
        fire_id = fire_incident['attributes'].get('LocalIncidentID') or fire_incident['attributes'].get(
            'UniqueFireIdentifier')
        fire_name = fire_incident['attributes'].get('IncidentName')
        print(f'{fire_name}: {fire_id}')
        ################################################################################################################
        # xlsx report
        now = dt.utcnow().strftime("%m%d%Y_%H%M%S")
        report_name = re.sub(r'\W+', '', '{0}_{1}'.format(fire_name, now)) + '.xlsx'
        report_path = BytesIO()
        # output report
        wb = openpyxl.load_workbook(report_template)

        # make overview sheet
        prepared_date = date.today()
        ws = wb.worksheets[0]
        t = ws.cell(1, 1).value
        ws.cell(1, 1, t.replace('[FireName]', fire_name))
        ws.cell(3, 1, prepared_date)
        t = ws.cell(4, 1).value
        ws.cell(4, 1, t.replace('[Analyst]', "R9 GIS Tech Center"))
        t = ws.cell(5, 1).value
        ws.cell(5, 1, t.replace('[Project#]', fire_id))
        # _10MiBuff
        header_fill = PatternFill(start_color='344C67', end_color='344C67', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        alternate_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

        ################################################################################################################
        reportstatus = 'successful'
        # ws_index = 0
        # for each facility layer populate facilities in correct sheet by index
        for pt_i, ptLayer in enumerate(config_settings['FACILITY_LAYERS']):
            try:
                fl_name = ptLayer['name']
                print(f'\n_________{fl_name}_________')
                print(f'processing {str(ptLayer)}')
                response['facilities'][fl_name] = 0
                # returns 2 layers  [fire, and 10 mi buffer intersections]
                boundary_intersects = get_intersect(service_url=ptLayer['url'], input_geom=fire_incident['geometry'],
                                                    token=token)
                buffer_intersects = get_intersect(service_url=ptLayer['url'], input_geom=fire_incident['buffer'],
                                                  token=token)
                lyrs = [boundary_intersects, buffer_intersects]
                # lyrs = fire_incident.intersect(ptLayer['url'], token)
                # NPL edge case, fix EPA_ID
                if ptLayer['name'] == 'NationalPriorityListPoint_R9_2019_R9':
                    # todo - adapt this edge case or move it
                    npl_polys = [x for x in config_settings['FACILITY_LAYERS'] if
                                 x['name'] == 'NationalPriorityListBoundaryTypes_R9_2020_R9'][0]
                    npl_poly_ids = \
                        FeatureLayer(npl_polys['url']).query(out_fields=['EPA_ID'], return_geometry=False, as_df=True)[
                            'EPA_ID'].to_list()
                    print(f'npl poly ids : {npl_poly_ids}')
                    print('processing npl pts')
                    for lyr in lyrs:
                        removeRows(target_lyr=lyr, deletes=npl_poly_ids, target_field='EPA_ID')
                for lyr_i, lyr in enumerate(lyrs):
                    sheet_index = ptLayer['sheet_index'] + lyr_i
                    populate_sheet(wb, sheet_index, ptLayer, fire_name, lyr)
                    feature_count = len(lyr.features) if lyr.features else 0
                    response['facilities'][fl_name] += feature_count
                    response['facilities']['Total'] += feature_count

                # success
                response['feature_geometry'] = fire_incident['buffer']
            except Exception as e:
                exc_type, exc_obj, exc_tb = sys.exc_info()
                fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
                out_msg = str(e) + " " + str(exc_type.__name__) + ' ' + str(fname) + ' ' + str(exc_tb.tb_lineno)
                print(out_msg)
                parent_sheet_index = pt_i + (2 * pt_i)

                # for i in [parent_sheet_index, parent_sheet_index + 1]:
                #     ws = wb.worksheets[i]
                #     error_msg_cell = ws.cell(5, 1)
                #     error_msg_cell.value = "There was an error getting intersecting facilities"
                reportstatus = 'error'
                print(f'Error getting features for {str(ptLayer)}')
                print(str(e))
                response['facilities'][ptLayer['name']] = [reportstatus, ptLayer['url'], str(e)]
                continue

        wb.save(report_path)
        print('_______total intersecting facilities {0} __________'.format(response['facilities']['Total']))
        print('report {} for fire id {}'.format(reportstatus, fire_id))
        # out_msg = str(json.dumps(response, indent=4))
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        out_msg = str(e) + " " + str(exc_type.__name__) + ' ' + str(fname) + ' ' + str(exc_tb.tb_lineno)
        out_msg = out_msg.replace('<', '_').replace('>', '_')
        print(out_msg)
        reportstatus = 'info' if fire_incident.fire is None else 'error'
        response['facilities']['Total'] = 0
        response['error_msg'] = out_msg

    finally:
        print("process execution took {} seconds".format(timer() - startmain))

        response['report_status'] = reportstatus
        response['RETRIEVED'] = int(dt.utcnow().timestamp() * 1000)
        # print(str(response))
        # if facilities: create feature
        file_dict = None
        if response['facilities']['Total'] > 0:
            if reportstatus == 'successful':
                file_dict = {
                    "contentType": "application/octet-stream",
                    "name": report_name,
                    "data": base64.b64encode(report_path.getvalue()).decode()
                }

        return response, file_dict


def upsert_msg(table_id, attributes, message_id=None, id_field='GlobalID'):
    gis = GIS('home')
    message_queue = gis.content.get(table_id)
    t = message_queue.tables[0]
    #     print(t)
    if message_id is not None:
        msg = t.query(where=f"{id_field} = '{message_id}'").features[0]
        for key in attributes:
            msg.attributes[key] = attributes[key]
        edit_features = {'updates': [msg]}
    else:
        msg = Feature(attributes=attributes)
        for key in attributes:
            msg.attributes[key] = attributes[key]
        edit_features = {'adds': [msg]}
    #     print(msg.attributes)
    return t.edit_features(**edit_features)


def update_unarchived_fires(irwin_fires, perimeter_fires, config_settings: dict, force_update=False):
    print('update unarchived')
    notifiable_fires_query = f"NotificationConfigurationID = '{config_settings['FIRE_CONFIG_ID']}'"
    notifiable_fires = load_features_json(url=config_settings['NOTIFIABLE_FEATURES'], where=notifiable_fires_query,
                                          auth_token=config_settings['TOKEN'], out_wkid=3857)
    if not notifiable_fires:
        print('no notifiable features matching query')
        return None
    for f in notifiable_fires:
        f['attributes']['Data'] = json.loads(f['attributes']['Data'])

    # get latest from IRWIN for existing
    irwin_ids_fallen_off = []
    fire_updates = []
    for fire in notifiable_fires:
        if fire['attributes'].get('Archived') is None:
            og_fire = deepcopy(fire)
            if fire['attributes'].get('Data').get('IRWINID') is not None:
                irwin_incidents = [x for x in irwin_fires if
                                   format_global(x['attributes']['IrwinID']).upper() == format_global(
                                       fire['attributes'].get('Data').get('IRWINID')).upper()]
                if len(irwin_incidents) > 0:
                    for incident in irwin_incidents:  # just use the last
                        fire['attributes']['Data']['percent_contained'] = incident['attributes']['PercentContained']
                        # fire['attributes']['Data']['_geometry'] = incident['geometry']
                        # fire['attributes']['Data']['_geometry']['spatialReference'] = {'wkid': 4326}
                        # fire['attributes']['Data']['counties'] = incident['attributes'].get('POOCounty')
                        fire['attributes']['Data']['acres'] = incident['attributes'].get('DailyAcres')
                # elif fire.perimeter_id is None:
                elif fire['attributes'].get('Data').get('perimeter_id', None) is None:
                    # if gone from irwin and no perimeter ID it can be archived
                    fire['attributes']['Archived'] = int(dt.utcnow().timestamp() * 1000)
                    # fire['attributes']['Display'] = 0
                else:
                    # irwin match doesn't exist, but perimeter id does - handled later
                    irwin_ids_fallen_off.append(fire.get('attributes').get('Data').get('IRWINID'))
                if force_update or fire != og_fire:
                    [fire_updates.remove(x) for x in fire_updates if
                     x['attributes'].get('GlobalID') == fire['attributes'].get('GlobalID')]
                    fire_updates.append(fire)
                    del og_fire

    ####################################################################################################################
    # see if a perimeter is available for fires with an irwin id
    # doesn't matter if upper or lower when querying globalid, but must have smooth braces
    unarchived_irwin_fires = [f['attributes'].get('Data').get('IRWINID')
                              for f in notifiable_fires if f['attributes'].get('Archived') is None
                              and f['attributes'].get('Data').get('IRWINID', None) is not None]
    print(f'{len(unarchived_irwin_fires)} unarchived fires')
    perimeters = [p for p in perimeter_fires if p['attributes']['IRWINID'] in unarchived_irwin_fires]
    for p in perimeters:
        # update fire with perimeter attribute info and geometry
        if p['attributes'].get('IRWINID', None) is not None:
            for fire in notifiable_fires:
                if fire['attributes'].get('Archived') is None and fire['attributes'].get('Data').get(
                        'IRWINID').lower() == p['attributes'].get('IRWINID').lower():
                    og_fire = deepcopy(fire)
                    fire['attributes']['Data']['perimeter_id'] = p['attributes'].get('GeometryID')
                    fire_name = p['attributes'].get('IncidentName') if p['attributes'].get('IncidentName') else ''
                    fire['attributes']['Data']['IncidentName'] = fire_name.upper()
                    # only update acres if missing (prefer get from irwin)
                    if fire['attributes']['Data'].get('acres', None) is None or fire['attributes'].get('Data').get(
                            'acres') == '':
                        fire['attributes']['Data']['acres'] = p['attributes'].get('GISAcres')
                    if force_update or fire != og_fire:
                        [fire_updates.remove(x) for x in fire_updates if
                         x['attributes'].get('GlobalID') == fire['attributes'].get('GlobalID')]
                        fire_updates.append(fire)
                        del og_fire

    ####################################################################################################################
    # see if irwin id is available for perimeter-generated fires without one
    perimeters_without_irwin = [i['attributes'].get('Data').get('perimeter_id') for i in notifiable_fires
                                if i['attributes'].get('Data').get('IRWINID', None) is None
                                and i['attributes'].get('Archived') is None
                                and i['attributes'].get('Data').get('perimeter_id', None) is not None]
    perimeters = [p for p in perimeters if p['attributes']['GeometryID'] in perimeters_without_irwin]
    # perim_updates = []
    for p in perimeters:
        for fire in notifiable_fires:
            if fire['attributes'].get('Archived') is None and fire['attributes'].get('Data').get(
                    'perimeter_id').lower() == p['attributes'].get('GeometryID').lower():
                og_fire = deepcopy(fire)
                fire['attributes']['Data']['IRWINID'] = p['attributes']['IRWINID']
                # fire['attributes']['Data']['_geometry'] = p['geometry']
                fire['attributes']['Data']['IncidentName'] = fire['attributes']['Name'] = p['attributes'][
                    'IncidentName'].upper()
                fire['attributes']['Data']['acres'] = p['attributes']['GISAcres']

                if force_update or fire != og_fire:
                    [fire_updates.remove(x) for x in fire_updates if
                     x['attributes'].get('GlobalID') == fire['attributes'].get('GlobalID')]
                    fire_updates.append(fire)
                    del og_fire
    # if perimeter is no longer found mark as archived if no irwin id
    for fire in notifiable_fires:
        # if not yet archived:
        if fire['attributes'].get('Archived') is None:
            fire_irwin_id = fire['attributes'].get('Data').get('IRWINID', None)
            fire_perim_id = fire['attributes'].get('Data').get('perimeter_id', None)
            if fire_irwin_id is None or fire_irwin_id in irwin_ids_fallen_off and not any(
                    x['attributes'].get('GeometryID') == fire_perim_id for x in perimeters):
                # f['attributes']['Archived'] = int(dt.now().timestamp()*1000)
                fire['attributes']['Archived'] = int(dt.utcnow().timestamp() * 1000)
                # fire['attributes']['Display'] = 0
                # update
                [fire_updates.remove(x) for x in fire_updates if
                 x['attributes'].get('GlobalID') == fire['attributes'].get('GlobalID')]
                fire_updates.append(fire)
    ####################################################################################################################
    if not fire_updates:
        print('no fire updates')
    for fire in fire_updates:
        print(f"Updating {fire.get('attributes').get('Name')} fire")
        fire_geom = fire['geometry'] if isinstance(fire['geometry'], Geometry) else Geometry(fire['geometry'])
        if fire['attributes']['Archived'] is not None:
            fire['attributes']['Display'] = 0
            incident_report = None
        else:
            incident_perim = fire.get('attributes').get('Data').get('perimeter_id')
            incident_results, incident_report = generate_fire_report(None, incident_perim) if incident_perim \
                else generate_fire_report(fire.get('attributes').get('Data').get('IRWINID'), None)

            if incident_results.get('feature_geometry', None) is not None:
                fire['geometry'] = incident_results.get('feature_geometry')
                del incident_results['feature_geometry']
            fire['attributes']['Data']['current_results'] = incident_results
            fire['attributes']['Retrieved'] = incident_results['RETRIEVED']
            # fire['attributes']['Display'] = incident_results.get('display', 0)
            # fire_geom = fire['geometry'] if isinstance(fire['geometry'], Geometry) else Geometry(fire['geometry'])

            fire['attributes']['Data']['counties'] = get_counties(fire_geom, config_settings['COUNTY_SERVICE'])
            tribes = get_tribes(fire_geom, config_settings['TRIBAL_LANDS'])
            fire['attributes']['Data']['tribes'] = tribes
            fire['attributes']['Display'] = display_fire(fire)
        # update
        fire['attributes']['Data'] = json.dumps(fire['attributes'].get('Data'))

        update_res = update_feature(Feature(geometry=fire['geometry'], attributes=fire['attributes']),
                                    feature_id=fire['attributes']['GlobalID'],
                                    target_url=config_settings['NOTIFIABLE_FEATURES'], attachment=incident_report)
        print(update_res)


def display_fire(fire_object):
    # returns 0, or 1 depending on criteria below
    if isinstance(fire_object, Feature):
        feature_attributes = fire_object._attributes
        feature_attributes['Data'] = json.loads(feature_attributes['Data'])
    else:
        feature_attributes = fire_object['attributes']
    minimal_acres = True if feature_attributes['Data'].get('acres') is None else feature_attributes['Data'].get('acres',
                                                                                                                0) < 10
    archived = True if feature_attributes.get('Archived', None) is not None else False
    contained = feature_attributes['Data'].get('percent_contained', 0) == 100
    if any([minimal_acres, archived, contained]):
        return 0
    tribes = feature_attributes['Data'].get('tribes')
    if tribes:
        return 1
    if feature_attributes['Data'].get('current_results'):
        active_rmp = feature_attributes['Data']['current_results']['facilities'].get('Active RMP Facilities', 0)
        npl_points = feature_attributes['Data']['current_results']['facilities'].get(
            'NationalPriorityListPoint_R9_2019_R9', 0)
        npl_polys = feature_attributes['Data']['current_results']['facilities'].get(
            'NationalPriorityListBoundaryTypes_R9_2020_R9', 0)
        facilities = active_rmp + npl_polys + npl_points
        if facilities:
            return 1
    return 0


def get_perimeters(perimeters_url, where=None, geometry=None, calc_centroids=False):
    perimeters = load_features_json(perimeters_url, where, geometry)
    if calc_centroids:
        for p in perimeters:
            p['attributes']['centroid'] = Polygon(
                {'spatialReference': {'latestWkid': 4326}, 'rings': p.get('geometry')['rings']}).centroid
    return perimeters


def get_irwin_info(irwin_url, where_statement=None, geometry_filter=None):
    irwin_incidents = load_features_json(irwin_url,
                                         where=where_statement, geometry=geometry_filter,
                                         order_by='ModifiedOnDateTime DESC')
    return irwin_incidents


@retry(stop=stop_after_attempt(3), after=after_log(get_logger(), logging.DEBUG))
def get_counties(input_geom, counties_url):
    geom_inter = intersects(input_geom,
                            input_geom['spatialReference'])
    counties = FeatureLayer(counties_url).query(geometry_filter=geom_inter,
                                                return_geometry=False)
    return ', '.join([c.attributes['NAME'].lower().title() for c in counties.features])


def get_tribes(input_geom, tribes_url):
    geom_inter = intersects(input_geom,
                            input_geom['spatialReference'])
    feats = FeatureLayer(tribes_url).query(geometry_filter=geom_inter,
                                           return_geometry=False)
    if not feats.features:
        return ''
    return ', '.join([c.attributes['Tribe_Name'].lower().title() for c in feats.features])


def get_extent(input_geom):
    print('get extent')
    if not isinstance(input_geom, Geometry):
        input_geom = Geometry(input_geom)
    spatial_ref = input_geom.spatial_reference.get('latestWkid') or input_geom.spatial_reference.get('wkid')
    if spatial_ref:
        return input_geom.extent + tuple([spatial_ref])
    return input_geom.extent


def compare_ids(a, b):
    if a is None or b is None:
        return False
    return format_global(a, False).upper() == format_global(b, False).upper()


def map_attributes(src_dict, dest_dict, attributes):
    for val in attributes:
        src, dst = val
        if type(val) == list:
            src = val[0]
            dst = val[1]
        dest_dict[dst] = src_dict[src]
    return


@retry(stop=stop_after_attempt(2), after=after_log(get_logger(), logging.DEBUG))
def update_custom_poi(token, config_settings: dict, id_field='GlobalID'):
    print("_______ update POIs ___________")
    # todo -- handle multiple poi's as a group
    # todo - pass in arrays of irwin and perimeter features
    all_pois = load_features_json(url=config_settings['CUSTOM_POI'], where="1=1", auth_token=token)
    notifiable_features = load_features_json(url=config_settings['NOTIFIABLE_FEATURES'], where="2=2",
                                             auth_token=token, out_wkid=3857)
    for poi_feature in all_pois:
        poi_id = poi_feature['attributes']['GlobalID']
        # prefer perimeter first
        poi_id = format_global(poi_id)
        poi_feature['geometry']['spatialReference'] = {'wkid': 102100, 'latestWkid': 3857}
        poi_buffer = buffer_miles(Geometry(poi_feature['geometry']))
        # Irwin Points
        irwin_where = f"IncidentTypeCategory = 'WF' AND PercentContained <> 100 AND DailyAcres >= 5"
        irwin_intersects = get_irwin_info(config_settings['IRWIN_SERVICE'], irwin_where, poi_buffer)
        # Perimeter Polygons
        perim_where = "GISAcres >= 5"
        perimeter_intersects = get_perimeters(config_settings['PERIMETER_SERVICE'], perim_where, poi_buffer)

        # update unarchived
        # existing notifiable features
        existing_feats = [x for x in notifiable_features if
                          compare_ids(x['attributes']['NotificationConfigurationID'], poi_id) and x['attributes'][
                              'Archived'] is None]

        # matching_perim = matching_irwin = []
        for feat in existing_feats:
            print(f'updating {feat["attributes"]}')
            feat['attributes']['Data'] = json.loads(feat['attributes']['Data']) if feat['attributes'].get('Data',
                                                                                                          False) else {}
            feat_perim_id = feat['attributes'].get('Data').get('perimeter_id')
            feat_irwin_id = feat['attributes'].get('Data').get('IRWINID')
            matching_perim = [p for p in perimeter_intersects if
                              compare_ids(p['attributes'].get('GeometryID'), feat_perim_id) or compare_ids(
                                  p['attributes'].get('IRWINID'), feat_irwin_id)]
            matching_irwin = [i for i in irwin_intersects if compare_ids(i['attributes'].get('IrwinID'), feat_irwin_id)]

            src_feat = matching_perim[0] if matching_perim else matching_irwin[0] if matching_irwin else None
            # now remove?
            if src_feat is not None:
                print(f'found matching fire feature "{src_feat["attributes"]["IncidentName"]}"')
                # attributes
                feat['attributes']['Data']['IRWINID'] = \
                    src_feat['attributes'].get('IrwinID', src_feat['attributes'].get('IRWINID', None))
                if src_feat['attributes'].get('GeometryID'):
                    feat['attributes']['Data']['perimeter_id'] = src_feat['attributes'].get('GeometryID')
                feat['attributes']['Data']['acres'] = src_feat['attributes'].get('GISAcres', src_feat['attributes'].get(
                    'DailyAcres', None))
                feat['attributes']['Data']['counties'] = get_counties(Geometry(feat['geometry']),
                                                                      config_settings['COUNTY_SERVICE'])
                feat['attributes']['Data']['tribes'] = get_tribes(Geometry(feat['geometry']),
                                                                  config_settings['TRIBAL_LANDS'])
                feat['geometry'] = buffer_miles(Geometry(src_feat['geometry']), in_wkid=4326)
                percent_contained = None
                archive = False
                # update
                # get irwin info no matter what
                if matching_irwin:
                    percent_contained = matching_irwin[0]['attributes']['PercentContained']
                    feat['attributes']['Data']['percent_contained'] = percent_contained
                    if percent_contained == 100:
                        archive = True
                if feat['attributes']['Data'].get('IRWINID') and not matching_irwin:
                    archive = True
            else:
                archive = True
            if archive:
                feat['attributes']['Archived'] = int(dt.utcnow().timestamp() * 1000)
            feat['attributes']['Retrieved'] = int(dt.utcnow().timestamp() * 1000)
            # update feature
            current_results = {}
            geom_extent = get_extent(feat['geometry'])
            current_results['FireBufferExtent'] = geom_extent
            feat['attributes']['Data']['current_results'] = json.dumps(current_results)
            feat['attributes']['Data'] = json.dumps(feat['attributes']['Data'])
            up_feat = Feature(geometry=feat['geometry'], attributes=feat['attributes'])
            update_res = update_feature(up_feat,
                                        config_settings['NOTIFIABLE_FEATURES'],
                                        feature_id=feat['attributes']['GlobalID'])
            print(update_res)
            [perimeter_intersects.remove(x) for x in matching_perim]
            [irwin_intersects.remove(x) for x in matching_irwin]

        # add new notifiable features
        perim_irwinids = [format_global(x['attributes']['IRWINID']).upper() for x in perimeter_intersects if
                          x['attributes'].get('IRWINID', None) is not None]
        incidents = perimeter_intersects + [i for i in irwin_intersects if
                                            not any(compare_ids(i['attributes']['IrwinID'], x) for x in perim_irwinids)]
        # new
        for incident in incidents:
            print(
                f"found incident for {poi_feature['attributes']['Name']} : {incident['attributes'].get('IncidentName')}")
            feature_attributes = {}
            feature_attributes['Data'] = {}
            irwin_id = feature_attributes['Data']['IRWINID'] = incident['attributes'].get('IrwinID',
                                                                                          incident['attributes'].get(
                                                                                              'IRWINID', None))
            perim_id = feature_attributes['Data']['perimeter_id'] = incident['attributes'].get('GeometryID', None)
            # get irwin discovery datetime if possible
            if perim_id and irwin_id:
                match_irwin = [i for i in irwin_intersects if compare_ids(i['attributes']['IrwinID'], irwin_id)]
                if match_irwin:
                    incident['attributes']['FireDiscoveryDateTime'] = match_irwin[0]['attributes'].get(
                        'FireDiscoveryDateTime')
            feature_attributes['Data']['counties'] = get_counties(Geometry(incident['geometry']))
            feature_attributes['Data']['tribes'] = get_tribes(Geometry(incident['geometry']),
                                                              config_settings['TRIBAL_LANDS'])
            feature_attributes['Data']['acres'] = incident['attributes'].get('GISAcres',
                                                                             incident['attributes'].get('DailyAcres',
                                                                                                        None))
            reported_date = incident['attributes'].get('FireDiscoveryDateTime',
                                                       incident['attributes'].get('CreateDate'))
            feature_attributes['Data']['reported_date'] = str(
                dt.fromtimestamp(reported_date / 1000)) if reported_date is not None else None

            feature_attributes['NotificationConfigurationID'] = format_global(poi_id, False)
            feature_attributes['Retrieved'] = int(dt.utcnow().timestamp() * 1000)
            feature_attributes['Name'] = incident['attributes']['IncidentName'].upper()
            incident_buffer = buffer_miles(Geometry(incident['geometry']), in_wkid=4326)
            feature_attributes['Data']['current_results'] = json.dumps({'FireBufferExtent': incident_buffer.extent})
            data = json.dumps(feature_attributes['Data'])
            feature_attributes['Data'] = data
            update_res = update_feature(Feature(geometry=incident_buffer, attributes=feature_attributes),
                                        target_url=config_settings['NOTIFIABLE_FEATURES'])
            print(update_res)
    print('done updating custom pois')
    return


def update_ipynb(input_file, agol_un, agol_pw, agol_id=None, item_properties=None, input_func=[]):
    if input_func is None:
        input_func = []
    gis = GIS(username=agol_un, password=agol_pw)
    py_script = open(input_file, 'r')
    py_script_contents = py_script.read()
    py_script.close()

    lines = py_script_contents.split('\n')
    if input_func:
        for func_file in input_func:
            import_line = [l for l in lines if func_file.split('/')[1].split('.')[0] in l and 'import' in l]
            if import_line:
                import_line = import_line[0]
                with open(func_file, 'r') as file:
                    file_contents = file.read()
                    utc_now = dt.utcnow().strftime('%x at %X UTC')
                    markdown = f"\n# %% [markdown] ##Updated {utc_now}\n"
                    edits = markdown + py_script_contents.replace(import_line, '\n' + file_contents + '\n')

                    with open(input_file, 'w') as py_script:
                        py_script.write(edits)

    basename = os.path.basename(input_file).split('.')[0]
    ipynb_output = input_file.replace('.py', '.ipynb')
    if os.path.basename(input_file).split('.')[1] == "py":
        input_py = jupytext.read(input_file)
        jupytext.write(input_py, ipynb_output)
    if not os.path.exists(ipynb_output):
        raise Exception('error generating ipynb')
    item = None
    print(f'ipynb: {ipynb_output}')
    print(f'basename: {basename}')
    # if agol_id then update existing
    if agol_id:
        item = gis.content.get(agol_id)
    else:
        search = gis.content.search(f"title:{basename}", item_type='Notebook')
        search = [x for x in search if x.title == basename]
        if search:
            item = search[0]
    if item:
        # update
        print(f'updating existing notebook {item.title}')
        item.update(data=ipynb_output, item_properties=item_properties)
    else:
        if not item_properties:
            item_properties = {'title': basename,
                               'type': 'Notebook',
                               'tags': f'{basename},notebook,autodeploy'}
        print('creating new notebook')
        item = gis.content.add(item_properties=item_properties, data=ipynb_output)
    return item
