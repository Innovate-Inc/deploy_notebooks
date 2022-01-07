
# install
#%%
pip install openpyxl

#%%
pip install slack_sdk

# %% [markdown]
## Markdowncell here xx
# end install

# import
#%%
import re
import os
import sys
import time
import openpyxl
from openpyxl.styles import PatternFill, Font
from openpyxl.utils.cell import get_column_letter
from datetime import date, timedelta, datetime as dt
from timeit import default_timer as timer
from arcgis.features import FeatureLayer, Feature, FeatureSet
from arcgis.geometry import Geometry, Polygon, Point, union
from arcgis.geometry.filters import intersects
from arcgis.geometry.functions import buffer, project
from io import BytesIO
from uuid import uuid4
import base64
from arcgis.gis import GIS
import urllib
import requests
import json
from copy import deepcopy
from tenacity import retry, stop_after_attempt, after_log
import logging
from slack_sdk.webhook import WebhookClient

#%%
import arcgis
arcgis.__version__

# import end

# config
#%%
FIRE_REPORT_SETTINGS = {
    'PERIMETER_SERVICE': 'https://services9.arcgis.com/RHVPKKiFTONKtxq3/ArcGIS/rest/services/USA_Wildfires_v1/FeatureServer/1',
    'IRWIN_SERVICE': 'https://services9.arcgis.com/RHVPKKiFTONKtxq3/ArcGIS/rest/services/USA_Wildfires_v1/FeatureServer/0',
    # 'BOUNDARIES_SERVICE': 'https://gispub.epa.gov/arcgis/rest/services/Region9/Boundaries_R9Administrative/MapServer/29',
    'BOUNDARIES_SERVICE': 'https://services.arcgis.com/P3ePLMYs2RVChkJx/ArcGIS/rest/services/USA_States_Generalized/FeatureServer/0',
    'COUNTY_SERVICE': 'https://services.arcgis.com/P3ePLMYs2RVChkJx/ArcGIS/rest/services/USA_Counties_Generalized/FeatureServer/0',
    # 'BUFFERED_PERIMETER_SERVICE': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/R9_Fire_Perimeter_Buffers/FeatureServer/0',
    # 'COP_LAYERS': 'USA_Wildfires_v1_3262;USA_Wildfires_v1_5448;PotentialPollutantAndSpillSources_6033;PotentialPollutantAndSpillSources_6033_14;PotentialPollutantAndSpillSources_6033_44;PotentialPollutantAndSpillSources_6033_62;PotentialPollutantAndSpillSources_6033_59;PotentialPollutantAndSpillSources_6033_58;PotentialPollutantAndSpillSources_6033_57;PotentialPollutantAndSpillSources_6033_19;PotentialPollutantAndSpillSources_6033_152;ER1702727_SafeDrinkingWater_9569;ER1702727_SafeDrinkingWater_9569_0;NationalPriorityListBoundaryTypes_R9_2020_R9_6264;ActiveSitesResponseepagov_2512;ActiveSitesResponseepagov_2512_0',
    # 'NOTIFIABLE_FEATURES': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/R9NotifiableFeatures/FeatureServer/0',
    'NOTIFIABLE_FEATURES': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/R9Notifiable/FeatureServer/0',
    # 'NOTEBOOK_ID': 'b9f4201a65ae44c2a3878cb563513234',
    'FIRE_CONFIG_ID': '7c7e8175-1aab-4092-9091-99af45148ab7',
    'TASK_TABLE_ID': '21b721732c3d4cf2bb1a0fe5fc4863eb',
    'CUSTOM_POI': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/R9NotificationCustomPoints/FeatureServer/0',
    'TRIBAL_LANDS': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/arcgis/rest/services/Tribal_Lands_R9_2020_BIA_BLM_EPA_Public_Layer_View/FeatureServer/0'
}

# Parameters
analyst = "R9 GIS Tech Center"
gis = GIS('home')
token = gis._con.token
# if fcntl module error, token may be bad...
# token = '0jiLJa6vnejxYEnsBnsioIp3B3SkOn_op3FcFzQU9SGOXf2MOWyB6VP5L4JkD19ZV8OnfnrVuVt6o9Hz8UDZUI8HkYOLw7vOALU3LYE8aVPmNSN00qYsCskJglPeBFthY3uImJMbpktiLIzacXJxcdv_YIYnX-HWcE9sIkyA3NVt4nfRNx5_zAaHPfXwHIK2Fu7_bN8ovd2jUiBA-XZNLUOJ1HCk-S9oW1Rif1YQr6U.'
# gis = GIS(token=token)
url = "https://hooks.slack.com/services/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"
webhook = WebhookClient(url)

logging.basicConfig(stream=sys.stderr, level=logging.DEBUG)
logger = logging.getLogger(__name__)

# config end

#%%
# functions
########################################################################################################################
@retry(stop=stop_after_attempt(3), after=after_log(logger, logging.DEBUG))
def load_feature_set(url, where, orderby=None, recordcount=None, fields="*", returngeometry='true'):
    fl = FeatureLayer(url)
    try:
        results = fl.query(where=where, return_geometry=returngeometry, out_fields=fields, orderby=orderby,
                           return_all_records=True)
        for f in results.features:
            f.geometry['spatialReference'] = results.spatial_reference
            f.geometry = Geometry(f.geometry)

        #         my_map.draw(results.features)
        print('load_feature_set' + " " + where)
        # print(results)
        return results.features
    #         return feature_set
    except Exception as e:
        print(e)
        return None


@retry(stop=stop_after_attempt(3), after=after_log(logger, logging.DEBUG))
def load_features_json(url, where=None, geometry=None, order_by=None, fields="*", attempts=2, auth_token=None,
                       out_wkid=4326):
    print(url)
    print(where)
    if geometry is not None:
        print('geometry')
    try:
        wkid = 4326
        # print('loading fs, attempt number 1')
        params = {'outFields': '*', 'f': 'json', 'inSR': wkid, 'ourSR': wkid,
                  'returnGeometry': True,
                  }
        if where:
            params["where"] = where
        else:
            params["where"] = "1=1"
        if geometry:
            params["geometry"] = geometry
            params['geometryType'] = 'esriGeometryPolygon'
            params['spatialRel'] = 'esriSpatialRelIntersects'
        if auth_token:
            params['token'] = auth_token

        data = urllib.parse.urlencode(params)
        if not url.endswith('/query'):
            url += '/query'
        request_response = requests.post(url, data=data,
                                         headers={'Content-Type': 'application/x-www-form-urlencoded'})
        if request_response.status_code != 200:
            raise Exception(f'request response status {request_response.status_code}')
        if request_response.json().get('error', None) is not None:
            raise Exception(request_response.json().get('error'))

        features = request_response.json().get('features', [])
        for f in features:
            if f.get('geometry', None) is None:
                print(f)
            f['geometry']['spatialReference'] = {'wkid': out_wkid}
        return features
    except Exception as e:
        print(e)
        print(url)
        attempts -= 1
        pass
        if attempts > 0:
            print(f'attempts left {attempts}')
            time.sleep(.1)
            load_features_json(url, where, geometry, order_by, fields, attempts)
        raise Exception(e)


@retry(stop=stop_after_attempt(3), after=after_log(logger, logging.DEBUG))
def update_feature(input_feature, target_url, feature_id=None, id_field='GlobalID', attachment=None):
    target_fl = FeatureLayer(target_url)
    attributes = {}
    for field in target_fl.properties['fields']:
        attributes[field['name']] = input_feature.attributes.get(field['name'], None)
    new_feat = Feature(geometry=input_feature.geometry, attributes=attributes)
    if feature_id is not None:
        feature_id = format_global(feature_id)
        query = f"{id_field} = '{feature_id}'"
        existing_feat = target_fl.query(where=query, out_fields='*')
        if len(existing_feat.features) > 0:
            feature_attributes = existing_feat.features[0].attributes
            new_feat.attributes['OBJECTID'] = feature_attributes['OBJECTID']
            new_feat.attributes['GlobalID'] = feature_attributes['GlobalID']
            edit_features = {'updates': [new_feat]}
        else:
            print('Could not find matching feature to update')
            new_feat.attributes.pop('OBJECTID')
            new_feat.attributes.pop('GlobalID')
            new_feat.attributes['GlobalID'] = str(uuid4())
            new_feat.attributes[id_field] = feature_id
            edit_features = {'adds': [new_feat]}
            # update_feature(input_feature, target_url, feature_id=None, id_field=id_field, attachment=attachment)
    else:
        new_feat.attributes.pop('OBJECTID')
        new_feat.attributes.pop('GlobalID')
        new_feat.attributes['GlobalID'] = str(uuid4())

        edit_features = {'adds': [new_feat]}
    if attachment is not None:
        attachment['parentGlobalId'] = format_global(new_feat.attributes['GlobalID'])
        attachment['globalId'] = str(uuid4())

        r = target_fl.edit_features(**edit_features, use_global_ids=True, attachments={"adds": [attachment]})
    else:
        r = target_fl.edit_features(**edit_features, use_global_ids=True)
    # print(r)
    return r


@retry(stop=stop_after_attempt(3), after=after_log(logger, logging.DEBUG))
def get_intersect(service_url, input_geom):
    try:
        if service_url[-6:] != '/query':
            service_url += '/query'
        print(service_url)
        geom_type = 'esriGeometryPoint' if input_geom.type == 'Point' else 'esriGeometryPolyline' if input_geom.type == 'Polyline' else 'esriGeometryPolygon'

        data = {'geometry': input_geom, 'f': 'json', 'outFields': '*', 'spatialRel': 'esriSpatialRelIntersects',
                'geometryType': geom_type, 'token': token}
        data = urllib.parse.urlencode(data)
        headers = {
            'content-type': 'application/x-www-form-urlencoded'
        }
        response = requests.post(service_url, data=data, headers=headers)
        results = response.json()
        if len(results.get('features')) > 0:
            print(len(results.get('features')))
            return FeatureSet(features=results.get('features'))
        return FeatureSet(features=[])
    except Exception as e:
        print('get_intersect error: ')
        print(e)
        return None


@retry(stop=stop_after_attempt(3), after=after_log(logger, logging.DEBUG))
def buffer_miles(geom, distance=10, unit=9030, wkid=3857, in_wkid=3857):
    buffer_geom = buffer([geom], in_sr=in_wkid, distances=distance,
                         unit=unit,
                         out_sr=wkid)[0]
    buffer_geom.spatialReference = {'wkid': 102100, 'latestWkid': 3857}
    return buffer_geom


def format_global(input_id, braces=True):
    # id = id.lower()
    if braces:
        if '{' not in input_id:
            input_id = '{' + '{0}'.format(input_id) + '}'
        return input_id
    else:
        return input_id.replace('{', '').replace('}', '')


class FireIncident(object):
    '''
    fire incident
    '''

    def __init__(self, irwinID=None, perimeterID=None):
        self.irwin_id = self.perimeter_id = None

        if irwinID and perimeterID:
            raise Exception("May not provide irwinID and perimeterID")
        if perimeterID:
            print('Using perimeter ID')
            self.perimeter_id = format_global(perimeterID)
        if irwinID:
            print('Using IRWIN ID')
            self.irwin_id = format_global(irwinID)
            # self.format_irwin_id()

        self.feature_id = self.irwin_id if irwinID else self.perimeter_id if perimeterID else None
        print('Incident - {0}'.format(self.feature_id))
        # raise if no id?
        self.data_source = None
        #         self.irwin_pts_url = 'https://services3.arcgis.com/T4QMspbfLg3qTGWY/arcgis/rest/services/Active_Fires/FeatureServer/0'
        #         self.perimeter_url = 'https://services3.arcgis.com/T4QMspbfLg3qTGWY/arcgis/rest/services/Public_Wildfire_Perimeters_View/FeatureServer/0'
        self.irwin_pts_url = 'https://services9.arcgis.com/RHVPKKiFTONKtxq3/ArcGIS/rest/services/USA_Wildfires_v1/FeatureServer/0'
        self.perimeter_url = 'https://services9.arcgis.com/RHVPKKiFTONKtxq3/ArcGIS/rest/services/USA_Wildfires_v1/FeatureServer/1'

        # load all things fire incident
        if self.download_fire() is None:
            return None
        self.fire_buffer = buffer_miles(self.fire.geometry, in_wkid=4326)
        self.parse_fire_fields()
        self.get_extent()

    def download_fire(self):
        '''
        Download the fire perimeter for the provided irwin id, or perimeter ID
        :return: returns path to in_memory fire perimeter, or None
        '''
        print('download_fire')

        if self.irwin_id:
            print('irwin fire')
            # do irwin id download
            irwin_id_field = 'IrwinID'
            irwin_where = "{0}='{1}'".format(irwin_id_field, self.irwin_id)
            irwin_orderby = '&orderByFields = FireDiscoveryDateTime + DESC'
            irwin_recordcount = '&resultRecordCount=1'
            fires = load_feature_set(self.irwin_pts_url
                                     , where=irwin_where
                                     , orderby=irwin_orderby
                                     , recordcount=irwin_recordcount)
            if fires:
                self.data_source = ['point', 'IRWIN Fire Incident Points']
                for fire in fires:
                    fire.attributes['IRWINID'] = fire.attributes['IrwinID']
                    fire.attributes['CreateDate'] = fire.attributes.get('FireDiscoveryDateTime', None)
                    fire.attributes['DataSource'] = self.data_source[1]
                    fire.attributes['FireID'] = self.feature_id
                    fire.attributes['RETRIEVED'] = int(dt.utcnow().timestamp() * 1000)
                    self.fire = fire

            else:
                print("Fire with that IRWIN ID not found")
                self.fire = None
                return None

        elif self.perimeter_id:
            print('perimeter fire')
            # do nifc perimeter id download
            perimeter_id_field = 'GeometryID'
            perim_where = "{0}='{1}'".format(perimeter_id_field, self.perimeter_id)
            perim_fields = '*'
            perim_orderby = 'CreateDate + DESC'
            perim_recordcount = 1
            fires = load_feature_set(self.perimeter_url
                                     , where=perim_where
                                     , orderby=perim_orderby
                                     , recordcount=perim_recordcount)

            if fires:
                self.data_source = ['polygon', 'Current Perimeters']
                for fire in fires:
                    fire.attributes['FireID'] = self.feature_id
                    fire.attributes['RETRIEVED'] = int(dt.utcnow().timestamp() * 1000)
                    fire.attributes['DataSource'] = self.data_source[1]
                    fire.attributes['IRWINID'] = fire.attributes['IRWINID']
                    fire.attributes['CreateDate'] = fire.attributes.get('CreateDate', None)
                    self.fire = fire

            else:
                print('Fire perimeter with that ID not found')
                self.fire = None
                return None

        print('downloaded from {0}'.format(self.data_source))

        print('download ********************** {} successful *********************'.format(self.feature_id))

        return self.fire

    # def buffer_fire(self, distance=10, unit=9030):
    #     print('buffer_fire')
    #     '''
    #     :return: returns string to temp fire buffer
    #     '''
    #     # unit code from https://resources.arcgis.com/en/help/arcobjects-cpp/componenthelp/index.html#/esriSRUnitType_Constants/000w00000042000000/
    #     print('buffering')
    #     #         print(self.fire.geometry)
    #
    #     self.fire_buffer = \
    #         buffer([self.fire.geometry], in_sr=self.fire.geometry.spatialReference['wkid'], distances=distance,
    #                unit=unit,
    #                out_sr=3857)[0]
    #     self.fire_buffer.spatialReference = {'wkid': 102100, 'latestWkid': 3857}

    def parse_fire_fields(self):
        print('parse_fire_fields')
        # Calculate added fields
        self.irwin_id = self.fire.attributes.get('IRWINID', None)
        #         self.PercentContained = self.fire.attributes.get('PercentContained', None)
        #         self.DailyAcres = self.fire.attributes.get('DailyAcres', None)
        self.incidentname = self.fire.attributes['IncidentName']
        print(self.incidentname)
        # print(self.fire.attributes)
        # creation_datetime = self.fire.attributes.get('CreateDate',
        #                                              self.fire.attributes.get('FireDiscoveryDateTime', 'No Date'))
        #
        # print(f'create date {creation_datetime}')
        # self.createdate = dt.fromtimestamp(creation_datetime / 1000).strftime('%m/%d/%Y')
        print('got date')
        self.gacc = self.fire.attributes['GACC']
        self.localincidentid = self.fire.attributes.get('LocalIncidentID',
                                                        self.fire.attributes.get('LocalIncidentIdentifier',
                                                                                 self.fire.attributes.get(
                                                                                     'UniqueFireIdentifier', None)))
        print(self.localincidentid)
        # tribal lands, counties
        # todo - get these from fire report settings
        # tribal_lands = 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/arcgis/rest/services/Tribal_Lands_R9_2020_BIA_BLM_EPA_Public_Layer_View/FeatureServer/0'
        # counties = 'https://services.arcgis.com/P3ePLMYs2RVChkJx/ArcGIS/rest/services/USA_Counties_Generalized/FeatureServer/0'
        # tribal_lands_fl = FeatureLayer(tribal_lands)
        # counties_fl = FeatureLayer(counties)
        # print('getting tribes and counties')
        #         print(self.fire_buffer)

        # fire_intersect = intersects(self.fire_buffer)
        print('got fire intersect geometry filter')
        #         print(fire_intersect)
        # tribal_intersects = tribal_lands_fl.query(geometry_filter=fire_intersect)
        # counties_intersects = counties_fl.query(geometry_filter=fire_intersect)
        # print('got tribes and counties')

        # self.fire.attributes['tribes'] = json.dumps([x.attributes['Tribe_Name'] for x in tribal_intersects.features])
        # self.fire.attributes['counties'] = json.dumps([x.attributes['NAME'] for x in counties_intersects.features])

        self.fire.attributes['acres'] = self.fire.attributes.get('DailyAcres',
                                                                 self.fire.attributes.get('GISAcres', None))

        if self.irwin_id:
            # update the fields below with irwin information if available
            # get irwin fields if irwinID
            print('getting irwin fields')
            irwin_id_field = 'IrwinID'
            irwin_where = "{0}='{1}'".format(irwin_id_field, self.irwin_id)
            irwin_orderby = 'FireDiscoveryDateTime + DESC'
            irwin_recordcount = 1

            matched_irwin_fires = load_feature_set(self.irwin_pts_url
                                          , where=irwin_where
                                          , orderby=irwin_orderby
                                          , recordcount=irwin_recordcount)
            from_irwin_fields = ['PercentContained', 'DailyAcres']
            if matched_irwin_fires:
                matched_irwin = matched_irwin_fires[0]
                print('found irwin point')
                for f in from_irwin_fields:
                    self.fire.attributes[f] = matched_irwin.attributes[f]

                creation_datetime = matched_irwin.attributes['FireDiscoveryDateTime']
                print(f'create date {creation_datetime}')
                self.fire.attributes['CreateDate'] = dt.fromtimestamp(creation_datetime / 1000).strftime('%m/%d/%Y')

        else:
            print('no irwin point found')

    def get_extent(self):
        if not self.fire_buffer:
            print('Must run buffer analysis first - buffer_fire')
            return None
        #         desc = arcpy.Describe(self.fire_buffer)
        #         extent = [desc.extent.XMin, desc.extent.YMin, desc.extent.XMax, desc.extent.YMax,
        #                   desc.spatialReference.factoryCode]
        self.buffer_extent = list(self.fire_buffer.extent)
        self.buffer_extent.append(self.fire_buffer.spatialReference['latestWkid'])
        print('Buffer Extent: {}'.format(str(self.buffer_extent)))

    def intersect(self, inputFC):
        if not self.data_source:
            print('no fire perimeter exists, download and buffer first')
            return None
        # set counts to zero
        total_count = 0
        perim_count = buffer_count = 0
        intersect_results = get_intersect(inputFC, self.fire.geometry)
        print('fire intersect done')
        perim_count = len(intersect_results.features)
        print(f'{perim_count} features intersecting fire geometry')
        total_count += perim_count

        for f in intersect_results.features:
            f.attributes['FireID'] = self.feature_id
            f.attributes['DataSource'] = ", ".join(self.data_source)
            # f.attributes['CreateDate'] = self.createdate
            f.attributes['IRWINID'] = self.irwin_id
            f.attributes['IncidentName'] = self.incidentname
            f.attributes['PercentContained'] = self.fire.attributes.get('PercentContained', None)
            f.attributes['DailyAcres'] = self.fire.attributes.get('DailyAcres', None)

        ########################################################################################################################
        # 10 mi buffer selection

        buffer_intersect_results = get_intersect(inputFC, self.fire_buffer)
        # remove overlap with fire intersection
        if intersect_results.features:
            flds = [x['name'] for x in buffer_intersect_results.fields]
            unique_field = 'OBJECTID' if 'OBJECTID' in flds else 'FID' if 'FID' in flds else 'GLOBALID' if 'GLOBALID' in flds else 'no unique field name known'
            fire_int_oids = [x.attributes[unique_field] for x in intersect_results.features]
            self.removeRows(target_lyr=buffer_intersect_results, deletes=fire_int_oids, target_field=unique_field)

        buffer_count = len(buffer_intersect_results.features)
        print(f'{buffer_count} features intersecting fire buffer')
        total_count += buffer_count

        # if buffer_count > 0:
        for f in intersect_results.features:
            f.attributes['FireID'] = self.feature_id
            f.attributes['DataSource'] = ", ".join(self.data_source)
            # f.attributes['CreateDate'] = self.createdate
            f.attributes['IRWINID'] = self.irwin_id
            f.attributes['IncidentName'] = self.incidentname
            f.attributes['PercentContained'] = self.fire.attributes.get('PercentContained', None)
            f.attributes['DailyAcres'] = self.fire.attributes.get('DailyAcres', None)

        return [intersect_results, buffer_intersect_results]

    def removeRows(self, target_lyr, deletes, target_field):
        # target_lyr.features = [x for x in target_lyr.features if x.attributes[target_field] not in deletes]
        for f in target_lyr.features:
            if f.attributes[target_field] in deletes:
                print(f'removing feature: {target_field}={f.attributes[target_field]}')
                target_lyr.features.remove(f)
        for f in target_lyr.features:
            if f.attributes[target_field] in deletes:
                print(f'removing feature: {target_field}={f.attributes[target_field]}')
                target_lyr.features.remove(f)
        return


def generate_fire_report(irwin_id, perimeter_id):
    if irwin_id:
        print("perimeter_id: {}".format(irwin_id))
    elif perimeter_id:
        print("perimeter_id: {}".format(perimeter_id))
    startmain = timer()
    reportstatus = 'not generated'
    try:
        # path to report template
        report_template_item = gis.content.get('6522d1d6db994710918200948cddca0a')
        report_template = report_template_item.download()
        target_data = 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/arcgis/rest/services/R9_Fire_Perimeter_Buffers/FeatureServer/0'
        pointLayerList = [
            # RMP
            {'name': 'Active RMP Facilities',
             'url': 'https://utility.arcgis.com/usrsvcs/servers/a9dda0a4ba0a433992ce3bdffd89d35a/rest/services/SharedServices/RMPFacilities/MapServer/0',
             'update_date': '03/01/2021'},
            # CA TierII
            {'name': 'CA Tier II',
             'url': 'https://utility.arcgis.com/usrsvcs/servers/f7e36ad5c73f4a19a24877d920a27c0a/rest/services/EPA_EPCRA/TierIIFacilities/MapServer/0',
             'update_date': '2017'},
            #             # SDWIS
            #             {'name': 'Safe_Drinking_Water_(SDWIS)_Region_9_V1',
            #             'url': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/SDWIS_Base/FeatureServer/0',
            #             'update_date': '2018'},
            # NPDES Wastewater
            {'name': 'NPDES Wastewater',
             'url': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/FRS_INTERESTS_NPDES/FeatureServer/0',
             'update_date': '2012'},
            # NPL Points
            # if this name changes, adjust npl point poly edge case below!
            {'name': 'NationalPriorityListPoint_R9_2019_R9',
             'url': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/R9_National_Priority_List_Points/FeatureServer/0',
             'update_date': '2019'},
            # NPL Polygons
            {'name': 'NationalPriorityListBoundaryTypes_R9_2020_R9',
             'url': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/R9_NPL_Site_Boundaries/FeatureServer/0',
             'update_date': '2019'},
            # FRP
            {'name': 'FRP Facilities',
             'url': 'https://services.arcgis.com/cJ9YHowT8TU7DUyn/ArcGIS/rest/services/Facility_Response_Plan_Sites_Region_9_V1_D/FeatureServer/0',
             'update_date': '2019'},
        ]

        response = {'facilities': {'total': 0}}
        if not irwin_id and not perimeter_id:
            return None
        if irwin_id:
            fire_incident = FireIncident(irwinID=irwin_id)
            response['IRWINID'] = fire_incident.irwin_id
        if perimeter_id:
            fire_incident = FireIncident(perimeterID=perimeter_id)
            response['perimeter_id'] = fire_incident.perimeter_id
            if fire_incident.irwin_id:
                response['IRWINID'] = fire_incident.irwin_id
        # Error if fire not found
        if not fire_incident.fire:
            raise Exception("fire not found")
        # return buffer extent
        response['FireBufferExtent'] = fire_incident.buffer_extent
        # response['tribes'] = fire_incident.fire.attributes['tribes']

        # fire feature id
        fire_id = fire_incident.feature_id
        fire_name = fire_incident.incidentname
        print(f'{fire_name}: {fire_id}')

        # xlsx report
        now = dt.utcnow().strftime("%m%d%Y_%H%M%S")
        report_name = re.sub(r'\W+', '', '{0}_{1}'.format(fire_incident.incidentname, now)) + '.xlsx'
        report_path = BytesIO()
        # output report
        wb = openpyxl.load_workbook(report_template)

        # make overview sheet
        prepared_date = date.today()
        ws = wb.worksheets[0]
        t = ws.cell(1, 1).value
        ws.cell(1, 1, t.replace('[FireName]', fire_name))
        ws.cell(3, 1, prepared_date)
        t = ws.cell(4, 1).value
        ws.cell(4, 1, t.replace('[Analyst]', analyst))
        t = ws.cell(5, 1).value
        ws.cell(5, 1, t.replace('[Project#]', fire_id))
        # _10MiBuff
        header_fill = PatternFill(start_color='344C67', end_color='344C67', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        alternate_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')

        ###############################
        reportstatus = 'successful'
        ws_index = 0
        for pt_i, ptLayer in enumerate(pointLayerList):
            try:
                # lyrs = facility_fire_intersect(ptLayer, irwin_id, fire_perimeter, fire_buffer)
                fl_name = ptLayer['name']
                print(f'\n_________{fl_name}_________')
                print(f'processing {str(ptLayer)}')
                response['facilities'][fl_name] = 0
                # returns 2 layers  [fire, and 10 mi buffer intersections]
                lyrs = fire_incident.intersect(ptLayer['url'])
                # NPL edge case, fix EPA_ID
                if ptLayer['name'] == 'NationalPriorityListPoint_R9_2019_R9':
                    npl_polys = [x for x in pointLayerList if x['name'] == 'NationalPriorityListBoundaryTypes_R9_2020_R9'][0]
                    npl_poly_ids = FeatureLayer(npl_polys['url']).query(out_fields=['EPA_ID'], return_geometry=False, as_df=True)['EPA_ID'].to_list()
                    print('processing npl pts')
                    for l in lyrs:
                        fire_incident.removeRows(target_lyr=l, deletes=npl_poly_ids, target_field='EPA_ID')
                for lyr in lyrs:
                    #                     print(f'lyr: {str(lyr)}')
                    ws_index += 1  # first sheet [0] is the summary sheet, then one sheet for each layer, and layer buffer
                    # order of Points list corresponds to order of sheets in workbook
                    ws = wb.worksheets[ws_index]
                    tab_title = ws.cell(1, 1).value
                    ws.cell(1, 1, tab_title.replace('[FireName]', fire_name))
                    date_cell = ws.cell(3, 1)
                    date_cell.value = f"Current as of: {ptLayer['update_date']}"
                    source_cell = ws.cell(4, 1)
                    source_cell.value = f"Data Source: {ptLayer['name']}"

                    if lyr.features:
                        feature_count = len(lyr.features)
                        response['facilities'][fl_name] += feature_count
                        response['facilities']['total'] += feature_count
                        unwanted_fields = ['OBJECTID', 'SHAPE', 'AUTOID']
                        fields = [f for f in lyr.fields if not any(y in f['name'].upper() for y in unwanted_fields)]

                        field_list = [f['name'] for f in fields]
                        field_alias = [f['alias'] for f in fields]

                        # reorder 'NAME' field to index 2
                        if 'NAME' in field_alias:
                            name_index = field_alias.index('NAME')
                            ualias = field_alias[name_index]
                            uname = field_list[name_index]
                            field_list.pop(name_index)
                            field_alias.pop(name_index)
                            field_list.insert(2, uname)
                            field_alias.insert(2, ualias)

                        header_row = 5  # not zero based
                        # add header - field aliases
                        for header_cell, v in enumerate(field_alias, 1):
                            _cell = ws.cell(header_row, header_cell, v)
                            _cell.fill = header_fill
                            _cell.font = header_font
                            column_width = 4
                            if len(v) > column_width:
                                column_width = len(v)
                                # print(str(len(v)) + '  cell:' +get_column_letter(header_cell))
                            ws.column_dimensions[get_column_letter(header_cell)].width = column_width + 3
                        # need to refactor column width to dictionary object. get max column width of data, maybe not to exceed 120
                        # make sure the column is as least as wide as the column heading
                        #
                        for row_index, row in enumerate(lyr.features, 6):  # start at row 6
                            for cell, f in enumerate(field_list):
                                _cell = ws.cell(row_index, cell + 1, row.attributes[f])
                                # column_width = ws.column_dimensions[get_column_letter(cell)].width
                                # Need to refactor column width to a dictionary object
                                # if len(row[cell]) > column_width:
                                #     column_width = len(row[cell])
                                if row_index % 2 != 0:
                                    _cell.fill = alternate_fill
                # success
                response['feature_geometry'] = fire_incident.fire_buffer
            except Exception as e:
                parent_sheet_index = pt_i + (2 * pt_i)
                for i in [parent_sheet_index, parent_sheet_index + 1]:
                    ws = wb.worksheets[i]
                    error_msg_cell = ws.cell(5, 1)
                    error_msg_cell.value = "There was an error getting intersecting facilities"
                reportstatus = 'error'
                print(f'Error getting features for {str(ptLayer)}')
                print(str(e))
                response['facilities'][ptLayer['name']] = [reportstatus, ptLayer['url'], str(e)]
                continue

        wb.save(report_path)
        print('_______total intersecting facilities {0} __________'.format(response['facilities']['total']))
        print('report {} for fire id {}'.format(reportstatus, fire_id))
        out_msg = str(json.dumps(response, indent=4))
    except Exception as e:
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        out_msg = str(e) + " " + str(exc_type.__name__) + ' ' + str(fname) + ' ' + str(exc_tb.tb_lineno)
        out_msg = out_msg.replace('<', '_').replace('>', '_')
        print(out_msg)
        reportstatus = 'info' if fire_incident.fire is None else 'error'
        response['facilities']['total'] = 0
        response['error_msg'] = out_msg

    finally:
        print("process execution took {} seconds".format(timer() - startmain))

        response['report_status'] = reportstatus
        response['RETRIEVED'] = int(dt.utcnow().timestamp() * 1000)
        # print(str(response))
        # if facilities: create feature
        file_dict = None
        if response['facilities']['total'] > 0:
            if reportstatus == 'successful':
                file_dict = {
                    "contentType": "application/octet-stream",
                    "name": report_name,
                    "data": base64.b64encode(report_path.getvalue()).decode()
                }

        return response, file_dict
        # udpate task notifications table
        # tblid = '21b721732c3d4cf2bb1a0fe5fc4863eb'
        # # todo - enable this?
        # upsert_msg(table_id=tblid, attributes={'Status': reportstatus,
        #                                        'TaskOutput': out_msg,
        #                                        'TaskName': 'Notifiable Fires Notebook 2.0',
        #                                        'Location': 'Region 9',
        #                                        'Computer': gis.users.me.username})


def upsert_msg(table_id, attributes, message_id=None, id_field='GlobalID'):
    message_queue = gis.content.get(table_id)
    t = message_queue.tables[0]
    #     print(t)
    if message_id is not None:
        msg = t.query(where=f"{id_field} = '{message_id}'").features[0]
        for key in attributes:
            msg.attributes[key] = attributes[key]
        edit_features = {'updates': [msg]}
    else:
        msg = Feature(attributes=attributes)
        for key in attributes:
            msg.attributes[key] = attributes[key]
        edit_features = edit_features = {'adds': [msg]}
    #     print(msg.attributes)
    return t.edit_features(**edit_features)


def update_unarchived_fires(irwin_fires, perimeter_fires, force_update=False):
    print('update unarchived')
    notifiable_fires_query = f"NotificationConfigurationID = '{FIRE_REPORT_SETTINGS['FIRE_CONFIG_ID']}'"
    notifiable_fires = load_features_json(url=FIRE_REPORT_SETTINGS['NOTIFIABLE_FEATURES'], where=notifiable_fires_query,
                                          auth_token=token, out_wkid=3857)
    if not notifiable_fires:
        print('no notifiable features matching query')
        return None
    for f in notifiable_fires:
        f['attributes']['Data'] = json.loads(f['attributes']['Data'])

    # get latest from IRWIN for existing
    irwin_ids_fallen_off = []
    fire_updates = []
    for fire in notifiable_fires:
        if fire['attributes'].get('Archived') is None:
            og_fire = deepcopy(fire)
            if fire['attributes'].get('Data').get('IRWINID') is not None:
                # irwin_where = f"IrwinID = '{fire['attributes'].get('Data').get('IRWINID')}'"
                # print(irwin_where)
                # irwin_incidents = get_irwin_info(where_statement=irwin_where)
                irwin_incidents = [x for x in irwin_fires if
                                   format_global(x['attributes']['IrwinID']).upper() == format_global(
                                       fire['attributes'].get('Data').get('IRWINID')).upper()]
                if len(irwin_incidents) > 0:
                    for incident in irwin_incidents:  # just use the last
                        fire['attributes']['Data']['percent_contained'] = incident['attributes']['PercentContained']
                        # fire['attributes']['Data']['_geometry'] = incident['geometry']
                        # fire['attributes']['Data']['_geometry']['spatialReference'] = {'wkid': 4326}
                        # fire['attributes']['Data']['counties'] = incident['attributes'].get('POOCounty')
                        fire['attributes']['Data']['acres'] = incident['attributes'].get('DailyAcres')
                # elif fire.perimeter_id is None:
                elif fire['attributes'].get('Data').get('perimeter_id', None) is None:
                    # if gone from irwin and no perimeter ID it can be archived
                    fire['attributes']['Archived'] = int(dt.utcnow().timestamp() * 1000)
                    # fire['attributes']['Display'] = 0
                else:
                    # irwin match doesn't exist, but perimeter id does - handled later
                    irwin_ids_fallen_off.append(fire.get('attributes').get('Data').get('IRWINID'))
                if force_update or fire != og_fire:
                    [fire_updates.remove(x) for x in fire_updates if
                     x['attributes'].get('GlobalID') == fire['attributes'].get('GlobalID')]
                    fire_updates.append(fire)
                    del og_fire

    ####################################################################################################################
    # see if a perimeter is available for fires with an irwin id
    # doesn't matter if upper or lower when querying globalid, but must have smooth braces
    unarchived_irwin_fires = [f['attributes'].get('Data').get('IRWINID')
                              for f in notifiable_fires if f['attributes'].get('Archived') is None
                              and f['attributes'].get('Data').get('IRWINID', None) is not None]
    print(f'{len(unarchived_irwin_fires)} unarchived fires')
    perimeters = [p for p in perimeter_fires if p['attributes']['IRWINID'] in unarchived_irwin_fires]
    for p in perimeters:
        # update fire with perimeter attribute info and geometry
        if p['attributes'].get('IRWINID', None) is not None:
            for fire in notifiable_fires:
                if fire['attributes'].get('Archived') is None and fire['attributes'].get('Data').get(
                    'IRWINID').lower() == p['attributes'].get('IRWINID').lower():
                    og_fire = deepcopy(fire)
                    fire['attributes']['Data']['perimeter_id'] = p['attributes'].get('GeometryID')
                    fire_name = p['attributes'].get('IncidentName') if p['attributes'].get('IncidentName') else ''
                    fire['attributes']['Data']['IncidentName'] = fire_name.upper()
                    # only update acres if missing (prefer get from irwin)
                    if fire['attributes']['Data'].get('acres', None) is None or fire['attributes'].get('Data').get(
                        'acres') == '':
                        fire['attributes']['Data']['acres'] = p['attributes'].get('GISAcres')
                    if force_update or fire != og_fire:
                        [fire_updates.remove(x) for x in fire_updates if
                         x['attributes'].get('GlobalID') == fire['attributes'].get('GlobalID')]
                        fire_updates.append(fire)
                        del og_fire

    ####################################################################################################################
    # see if irwin id is available for perimeter-generated fires without one
    perimeters_without_irwin = [i['attributes'].get('Data').get('perimeter_id') for i in notifiable_fires
                                if i['attributes'].get('Data').get('IRWINID', None) is None
                                and i['attributes'].get('Archived') is None
                                and i['attributes'].get('Data').get('perimeter_id', None) is not None]
    perimeters = [p for p in perimeters if p['attributes']['GeometryID'] in perimeters_without_irwin]
    # perim_updates = []
    for p in perimeters:
        for fire in notifiable_fires:
            if fire['attributes'].get('Archived') is None and fire['attributes'].get('Data').get(
                'perimeter_id').lower() == p['attributes'].get('GeometryID').lower():
                og_fire = deepcopy(fire)
                fire['attributes']['Data']['IRWINID'] = p['attributes']['IRWINID']
                # fire['attributes']['Data']['_geometry'] = p['geometry']
                fire['attributes']['Data']['IncidentName'] = fire['attributes']['Name'] = p['attributes'][
                    'IncidentName'].upper()
                fire['attributes']['Data']['acres'] = p['attributes']['GISAcres']

                if force_update or fire != og_fire:
                    [fire_updates.remove(x) for x in fire_updates if
                     x['attributes'].get('GlobalID') == fire['attributes'].get('GlobalID')]
                    fire_updates.append(fire)
                    del og_fire
    # if perimeter is no longer found mark as archived if no irwin id
    for fire in notifiable_fires:
        # if not yet archived:
        if fire['attributes'].get('Archived') is None:
            fire_irwin_id = fire['attributes'].get('Data').get('IRWINID', None)
            fire_perim_id = fire['attributes'].get('Data').get('perimeter_id', None)
            if fire_irwin_id is None or fire_irwin_id in irwin_ids_fallen_off and not any(
                x['attributes'].get('GeometryID') == fire_perim_id for x in perimeters):
                # f['attributes']['Archived'] = int(dt.now().timestamp()*1000)
                fire['attributes']['Archived'] = int(dt.utcnow().timestamp() * 1000)
                # fire['attributes']['Display'] = 0
                # update
                [fire_updates.remove(x) for x in fire_updates if
                 x['attributes'].get('GlobalID') == fire['attributes'].get('GlobalID')]
                fire_updates.append(fire)
    ####################################################################################################################
    if not fire_updates:
        print('no fire updates')
    for fire in fire_updates:
        print(f"Updating {fire.get('attributes').get('Name')} fire")
        fire_geom = fire['geometry'] if isinstance(fire['geometry'], Geometry) else Geometry(fire['geometry'])
        if fire['attributes']['Archived'] is not None:
            fire['attributes']['Display'] = 0
            incident_report = None
        else:
            incident_perim = fire.get('attributes').get('Data').get('perimeter_id')
            incident_results, incident_report = generate_fire_report(None, incident_perim) if incident_perim \
                else generate_fire_report(fire.get('attributes').get('Data').get('IRWINID'), None)

            if incident_results.get('feature_geometry', None) is not None:
                fire['geometry'] = incident_results.get('feature_geometry')
                del incident_results['feature_geometry']
            fire['attributes']['Data']['current_results'] = incident_results
            fire['attributes']['Retrieved'] = incident_results['RETRIEVED']
            # fire['attributes']['Display'] = incident_results.get('display', 0)
            # fire_geom = fire['geometry'] if isinstance(fire['geometry'], Geometry) else Geometry(fire['geometry'])

            fire['attributes']['Data']['counties'] = get_counties(fire_geom)
            tribes = get_tribes(fire_geom)
            fire['attributes']['Data']['tribes'] = tribes
            fire['attributes']['Display'] = display_fire(fire)
        # update
        fire['attributes']['Data'] = json.dumps(fire['attributes'].get('Data'))

        update_res = update_feature(Feature(geometry=fire['geometry'], attributes=fire['attributes']),
                                    feature_id=fire['attributes']['GlobalID'],
                                    target_url=FIRE_REPORT_SETTINGS['NOTIFIABLE_FEATURES'], attachment=incident_report)
        print(update_res)

def display_fire(fire_object):
    # returns 0, or 1 depending on criteria below
    if isinstance(fire_object, Feature):
        feature_attributes = fire_object._attributes
        feature_attributes['Data'] = json.loads(feature_attributes['Data'])
    else:
        feature_attributes = fire_object['attributes']
    minimal_acres = True if feature_attributes['Data'].get('acres') is None else feature_attributes['Data'].get('acres', 0) < 10
    archived = True if feature_attributes.get('Archived', None) is not None else False
    contained = feature_attributes['Data'].get('percent_contained', 0) == 100
    if any([minimal_acres, archived, contained]):
        return 0
    tribes = feature_attributes['Data'].get('tribes')
    if tribes:
        return 1
    if feature_attributes['Data'].get('current_results'):
        active_rmp = feature_attributes['Data']['current_results']['facilities'].get('Active RMP Facilities', 0)
        npl_points = feature_attributes['Data']['current_results']['facilities'].get('NationalPriorityListPoint_R9_2019_R9', 0)
        npl_polys = feature_attributes['Data']['current_results']['facilities'].get('NationalPriorityListBoundaryTypes_R9_2020_R9', 0)
        facilities = active_rmp+npl_polys+npl_points
        if facilities:
            return 1
    return 0

def get_perimeters(where=None, geometry=None, calc_centroids=False):
    perimeters = load_features_json(FIRE_REPORT_SETTINGS['PERIMETER_SERVICE'], where, geometry)
    if calc_centroids:
        for p in perimeters:
            p['attributes']['centroid'] = Polygon(
                {'spatialReference': {'latestWkid': 4326}, 'rings': p.get('geometry')['rings']}).centroid
    return perimeters


def get_irwin_info(where_statement=None, geometry_filter=None):
    irwin_incidents = load_features_json(FIRE_REPORT_SETTINGS['IRWIN_SERVICE'],
                                         where=where_statement, geometry=geometry_filter,
                                         order_by='ModifiedOnDateTime DESC')
    return irwin_incidents


@retry(stop=stop_after_attempt(3), after=after_log(logger, logging.DEBUG))
def get_counties(input_geom):
    geom_inter = intersects(input_geom,
                            input_geom['spatialReference'])
    counties = FeatureLayer(FIRE_REPORT_SETTINGS['COUNTY_SERVICE']).query(geometry_filter=geom_inter,
                                                                          return_geometry=False)
    return ', '.join([c.attributes['NAME'].lower().title() for c in counties.features])

def get_tribes(input_geom):
    geom_inter = intersects(input_geom,
                            input_geom['spatialReference'])
    feats = FeatureLayer(FIRE_REPORT_SETTINGS['TRIBAL_LANDS']).query(geometry_filter=geom_inter,
                                                                          return_geometry=False)
    if not feats.features:
        return ''
    return ', '.join([c.attributes['Tribe_Name'].lower().title() for c in feats.features])


def get_spatial_intersect(input_geom, url):
    geom_inter = intersects(input_geom,
                            input_geom['spatialReference'])
    results = FeatureLayer(url).query(geometry_filter=geom_inter)
    return results

def get_extent(input):
    print('get extent')
    if isinstance(input, Geometry):
        ext = input.extent
        # default 3857
        return input.extent+tuple([input.spatial_reference.get('latestWkid', input.spatial_reference.get('wkid', 3857))])
    else:
        g = Geometry(input)
        return g.extent+tuple([g.spatial_reference.get('latestWkid', g.spatial_reference.get('wkid', 3857))])


def compare_ids(a, b):
    if a is None or b is None:
        return False
    return format_global(a, False).upper() == format_global(b, False).upper()


def map_attributes(src_dict, dest_dict, attributes):
    for val in attributes:
        src, dst = val
        if type(val) == list:
            src = val[0]
            dst = val[1]
        dest_dict[dst] = src_dict[src]
    return


@retry(stop=stop_after_attempt(2), after=after_log(logger, logging.DEBUG))
def update_custom_poi(auth_token, id_field='GlobalID'):
    print("_______ update POIs ___________")
    # todo -- handle multiple poi's as a group
    # todo - pass in arrays of irwin and perimeter features
    all_pois = load_features_json(url=FIRE_REPORT_SETTINGS['CUSTOM_POI'], where="1=1", auth_token=auth_token)
    notifiable_features = load_features_json(url=FIRE_REPORT_SETTINGS['NOTIFIABLE_FEATURES'], where="2=2",
                                             auth_token=token, out_wkid=3857)
    for poi_feature in all_pois:
        poi_id = poi_feature['attributes']['GlobalID']
        # prefer perimeter first
        poi_id = format_global(poi_id)
        poi_feature['geometry']['spatialReference'] = {'wkid': 102100, 'latestWkid': 3857}
        poi_buffer = buffer_miles(Geometry(poi_feature['geometry']))
        # Irwin Points
        irwin_where = f"IncidentTypeCategory = 'WF' AND PercentContained <> 100 AND DailyAcres >= 5"
        irwin_intersects = get_irwin_info(irwin_where, poi_buffer)
        # Perimeter Polygons
        perim_where = "GISAcres >= 5"
        perimeter_intersects = get_perimeters(perim_where, poi_buffer)

        # update unarchived
        # existing notifiable features
        existing_feats = [x for x in notifiable_features if
                          compare_ids(x['attributes']['NotificationConfigurationID'], poi_id) and x['attributes']['Archived'] is None]

        # matching_perim = matching_irwin = []
        for feat in existing_feats:
            print(f'updating {feat["attributes"]}')
            feat['attributes']['Data'] = json.loads(feat['attributes']['Data']) if feat['attributes'].get('Data',
                                                                                                          False) else {}
            feat_perim_id = feat['attributes'].get('Data').get('perimeter_id')
            feat_irwin_id = feat['attributes'].get('Data').get('IRWINID')
            matching_perim = [p for p in perimeter_intersects if
                              compare_ids(p['attributes'].get('GeometryID'), feat_perim_id) or compare_ids(p['attributes'].get('IRWINID'), feat_irwin_id)]
            matching_irwin = [i for i in irwin_intersects if compare_ids(i['attributes'].get('IrwinID'), feat_irwin_id)]

            src_feat = matching_perim[0] if matching_perim else matching_irwin[0] if matching_irwin else None
            # now remove?
            if src_feat is not None:
                print(f'found matching fire feature "{src_feat["attributes"]["IncidentName"]}"')
                # attributes
                feat['attributes']['Data']['IRWINID'] = \
                    src_feat['attributes'].get('IrwinID', src_feat['attributes'].get('IRWINID', None))
                if src_feat['attributes'].get('GeometryID'):
                    feat['attributes']['Data']['perimeter_id'] = src_feat['attributes'].get('GeometryID')
                feat['attributes']['Data']['acres'] = src_feat['attributes'].get('GISAcres', src_feat['attributes'].get(
                    'DailyAcres', None))
                # reported_date = src_feat['attributes'].get('CreateDate',
                #                                            src_feat['attributes'].get('FireDiscoveryDateTime'))
                # feat['attributes']['Data']['reported_date'] = str(
                #     dt.fromtimestamp(reported_date / 1000)) if reported_date is not None else None
                feat['attributes']['Data']['counties'] = get_counties(Geometry(feat['geometry']))
                feat['attributes']['Data']['tribes'] = get_tribes(Geometry(feat['geometry']))
                feat['geometry'] = buffer_miles(Geometry(src_feat['geometry']), in_wkid=4326)
                # feat['attributes']['Data']['current_results'] = {}
                # feat['attributes']['Data']['current_results']['FireBufferExtent'] = feat['geometry']
                percent_contained = None
                archive = False
                # update
                # get irwin info no matter what
                if matching_irwin:
                    percent_contained = matching_irwin[0]['attributes']['PercentContained']
                    feat['attributes']['Data']['percent_contained'] = percent_contained
                    # todo - don't need to include this since querying where <>100 anyway?
                    if percent_contained == 100:
                        archive = True
            else:
                archive = True
            if archive:
                feat['attributes']['Archived'] = int(dt.utcnow().timestamp() * 1000)
            feat['attributes']['Retrieved'] = int(dt.utcnow().timestamp() * 1000)
            # update feature
            current_results = {}
            geom_extent = get_extent(feat['geometry'])
            current_results['FireBufferExtent'] = geom_extent
            feat['attributes']['Data']['current_results'] = json.dumps(current_results)
            feat['attributes']['Data'] = json.dumps(feat['attributes']['Data'])
            up_feat = Feature(geometry=feat['geometry'], attributes=feat['attributes'])
            update_res = update_feature(up_feat,
                                        FIRE_REPORT_SETTINGS['NOTIFIABLE_FEATURES'],
                                        feature_id=feat['attributes']['GlobalID'])
            print(update_res)
            [perimeter_intersects.remove(x) for x in matching_perim]
            [irwin_intersects.remove(x) for x in matching_irwin]

        # add new notifiable features
        # match_perim_irwinids = [format_global(x['attributes']['IRWINID']).upper() for x in matching_perim if x['attributes'].get('IRWINID', None) is not None]
        perim_irwinids = [format_global(x['attributes']['IRWINID']).upper() for x in perimeter_intersects if
                          x['attributes'].get('IRWINID', None) is not None]
        # incidents = irwin_intersects+[x for x in perimeter_intersects if x['attributes']['IRWINID'] not in irwinids]
        # incidents = [p for p in perimeter_intersects if p not in matching_perim]
        # incidents += [x for x in irwin_intersects if format_global(x['attributes']['IrwinID']).upper() not in match_perim_irwinids
        #              and format_global(x['attributes']['IrwinID']).upper() not in matching_irwin and format_global(x['attributes']['IrwinID']).upper() not in perim_irwinids]
        incidents = perimeter_intersects + [i for i in irwin_intersects if
                                            not any(compare_ids(i['attributes']['IrwinID'], x) for x in perim_irwinids)]
        # new
        for incident in incidents:
            print(f"found incident for {poi_feature['attributes']['Name']} : {incident['attributes'].get('IncidentName')}")
            feature_attributes = {}
            # feature_attributes['GlobalID'] = \
            #     incident['attributes'].get('GeometryID', incident['attributes'].get('IrwinID'))
            # incident['attributes']['Data'] = incident['attributes']
            feature_attributes['Data'] = {}
            irwin_id = feature_attributes['Data']['IRWINID'] = incident['attributes'].get('IrwinID',
                                                                                          incident['attributes'].get(
                                                                                              'IRWINID', None))
            perim_id = feature_attributes['Data']['perimeter_id'] = incident['attributes'].get('GeometryID', None)
            #get irwin discovery datetime if possible
            if perim_id and irwin_id:
                match_irwin = [i for i in irwin_intersects if compare_ids(i['attributes']['IrwinID'], irwin_id)]
                if match_irwin:
                    incident['attributes']['FireDiscoveryDateTime'] = match_irwin[0]['attributes'].get('FireDiscoveryDateTime')
            # feature_attributes['Data']['IRWINID'] = irwin_id

            # feature_attributes['Data']['perimeter_id'] = perim_id
            feature_attributes['Data']['counties'] = get_counties(Geometry(incident['geometry']))
            feature_attributes['Data']['tribes'] = get_tribes(Geometry(incident['geometry']))
            feature_attributes['Data']['acres'] = incident['attributes'].get('GISAcres',
                                                                             incident['attributes'].get('DailyAcres',
                                                                                                        None))
            reported_date = incident['attributes'].get('FireDiscoveryDateTime', incident['attributes'].get('CreateDate'))
            feature_attributes['Data']['reported_date'] = str(
                dt.fromtimestamp(reported_date / 1000)) if reported_date is not None else None

            feature_attributes['NotificationConfigurationID'] = format_global(poi_id, False)
            feature_attributes['Retrieved'] = int(dt.utcnow().timestamp() * 1000)
            feature_attributes['Name'] = incident['attributes']['IncidentName'].upper()
            incident_buffer = buffer_miles(Geometry(incident['geometry']), in_wkid=4326)
            # feature_attributes['Data']['FireBufferExtent'] = Geometry(incident_buffer).extent
            # feature_attributes['Data']['current_results'] = {}
            #             feature_attributes['Data']['current_results'] = json.dumps({'FireBufferExtent': Geometry(incident_buffer).extent})
            feature_attributes['Data']['current_results'] = json.dumps({'FireBufferExtent': incident_buffer.extent})
            data = json.dumps(feature_attributes['Data'])
            feature_attributes['Data'] = data
            update_res = update_feature(Feature(geometry=incident_buffer, attributes=feature_attributes),
                                        target_url=FIRE_REPORT_SETTINGS['NOTIFIABLE_FEATURES'])
            print(update_res)
    print('done updating custom pois')
    return


# def main(days=1):
@retry(stop=stop_after_attempt(2), after=after_log(logger, logging.DEBUG))
def main():
    try:
        # print(f'main, days={days}')
        ####################################################################################################################

        notifiable_fires_query = f"NotificationConfigurationID = '{FIRE_REPORT_SETTINGS['FIRE_CONFIG_ID']}'"
        fire_features = load_features_json(url=FIRE_REPORT_SETTINGS['NOTIFIABLE_FEATURES'],
                                           where=notifiable_fires_query,
                                           auth_token=token)
        for f in fire_features:
            f['attributes']['Data'] = json.loads(f['attributes']['Data'])
        # get R9 CONUS boundary
        conus_where = "STATE_ABBR='CA' OR STATE_ABBR='AZ' OR STATE_ABBR='NV'"
        r9_features = load_features_json(FIRE_REPORT_SETTINGS['BOUNDARIES_SERVICE'], where=conus_where)
        # union into single poly
#         r9_geom = union(3857, [x['geometry'] for x in r9_features])
#         r9_geom['spatialReference'] = {'wkid': 102100, 'latestWkid': 3857}
        # irwin_where = f"IncidentTypeCategory = 'WF' AND FireDiscoveryDateTime > CURRENT_TIMESTAMP - {days}"
        irwin_where = "IncidentTypeCategory = 'WF'"
        irwin_incidents = [item for feat in r9_features for item in get_irwin_info(where_statement=irwin_where, geometry_filter=feat['geometry'])]
        perimeter_incidents = [item for feat in r9_features for item in get_perimeters(geometry=feat['geometry'], calc_centroids=True)]
        # update existing
        update_unarchived_fires(irwin_incidents, perimeter_incidents, force_update=True)
        update_custom_poi(auth_token=token)

        # get fires we know about and notified already
        # known_irwin_ids = list(Fire.objects.filter(irwin_id__isnull=False).values_list('irwin_id', flat=True))
        known_irwin_ids = [x['attributes'].get('Data').get('IRWINID') for x in fire_features if
                           x['attributes'].get('Data', None).get('IRWINID', None) is not None]
        # known_perimeter_ids = list(
        #     Fire.objects.filter(perimeter_id__isnull=False).values_list('perimeter_id', flat=True))
        known_perimeter_ids = [x['attributes'].get('Data').get('perimeter_id') for x in fire_features if
                               x['attributes'].get('Data').get('perimeter_id', None) is not None]
        ################################################################################################################
        new_incidents = []
        # prevent duplicate irwin points and just get most recently modified
        for x in irwin_incidents:
            if x['attributes']['IrwinID'] not in known_irwin_ids and \
                not any(y for y in new_incidents if y['IRWINID'] == x['attributes']['IrwinID']):
                new_incidents.append({'IncidentName': x['attributes']['IncidentName'].upper(),
                                      'IRWINID': x['attributes']['IrwinID'],
                                      'reported_date': str(dt.fromtimestamp(
                                          x['attributes']['FireDiscoveryDateTime'] / 1000)),
                                      # '_geometry': x['geometry'],
                                      # 'geometry': x.get('geometry'),
                                      'percent_contained': x['attributes']['PercentContained'],
                                      # 'counties': x['attributes']['POOCounty'],
                                      'acres': x['attributes']['DailyAcres']
                                      })

        # get perimeters for irwin incidents and update new incident record with perimeter ID
        perimeters = []
        incidents_with_irwinid = [i['IRWINID'] for i in new_incidents if i.get('IRWINID', False)]
        if len(incidents_with_irwinid) > 0:
            matching_perims = [p for p in perimeter_incidents if p['attributes']['IRWINID'] in incidents_with_irwinid]
            for p in matching_perims:
                index = \
                    [i for i, x in enumerate(new_incidents) if
                     f"{x.get('IRWINID').upper()}" == p.get('attributes').get('IRWINID').upper()][
                        0]
                new_incidents[index]['perimeter_id'] = p['attributes']['GeometryID']
                # capture acres if irwin DailyAcres is null
                if new_incidents[index]['acres'] is None:
                    new_incidents[index]['acres'] = p['attributes']['GISAcres']

        perims_missing_irwin = [p for p in perimeter_incidents if p['attributes'].get('IRWINID') is None]
        perimeters += perims_missing_irwin

        # merge new_indcidents and perimeters
        new_incidents += [{'IncidentName': p['attributes']['IncidentName'].upper(),
                           'perimeter_id': p['attributes']['GeometryID'],
                           'reported_date': str(dt.fromtimestamp(p['attributes']['CreateDate'] / 1000)),
                           # '_geometry': json.dumps(p['geometry']),
                           'geometry': p.get('geometry'),
                           'acres': p['attributes']['GISAcres']} for p in perimeters
                          if p['attributes']['GeometryID'] not in known_perimeter_ids]
        ######################################################################################################################
        # log newly found fires
        new_fires = []
        for incident in new_incidents:
            if incident.get('acres', None) is None or incident.get('acres', 0) < 10:
                continue
            incident_irwin = incident.get('IRWINID', None)
            incident_perim = incident.get('perimeter_id', None)
            incident_results, incident_report = generate_fire_report(incident_irwin, incident_perim)
            # create new feature
            feat = Feature(attributes={})
            feat.geometry = incident_results.get('feature_geometry')
            del incident_results['feature_geometry']
            feat.attributes['Name'] = incident.get('IncidentName').upper()
            feat.attributes['Retrieved'] = incident_results['RETRIEVED']
            incident['current_results'] = incident_results
            incident['counties'] = get_counties(feat.geometry)
            tribes = get_tribes(feat.geometry)
            incident['tribes'] = tribes
            feat.attributes['Data'] = json.dumps(incident)
            feat.attributes['Display'] = display_fire(feat)
            feat.attributes['NotificationConfigurationID'] = FIRE_REPORT_SETTINGS['FIRE_CONFIG_ID']

            # new_fires.append(feat)
            created_ft = update_feature(feat, target_url=FIRE_REPORT_SETTINGS['NOTIFIABLE_FEATURES'],
                                        attachment=incident_report)
            print(f'Created {feat.attributes["Name"]} feature')
            print(created_ft)
        print('main done')
        # response = webhook.send(text="R9 Fire Notifications Notebook Run: SUCCESS")
        return
    except Exception as e:
        print(e)
        exc_type, exc_obj, exc_tb = sys.exc_info()
        fname = os.path.split(exc_tb.tb_frame.f_code.co_filename)[1]
        error_msg = ", ".join([str(exc_type), str(fname), str(exc_tb.tb_lineno)])
        print(error_msg)
        response = webhook.send(text="R9 Fire Notifications Notebook Run: FAILED")
        response = webhook.send(text=error_msg)
        raise Exception(e)

# functions end

#%%
main()


# generate_fire_report(None, "{A6EC388F-A752-4A23-9721-62E8DF4F3BEE}") # coldwater perimeter
# generate_fire_report(None, "{BD4DF34B-1CD1-4816-AF97-19998B5848CB}") # mccash perimeter
# update_custom_poi(token)
# update_unarchived_fires()



